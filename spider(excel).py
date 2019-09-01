from selenium import webdriver
import time
import json
import openpyxl
from openpyxl.utils import get_column_letter

def main():
    # 因为我太菜了不懂怎么爬那个表只能先用selenium了
    br = webdriver.Chrome()
    url = "http://bond.sse.com.cn/disclosure/info/tb/#"
    br.get(url)
    time.sleep(1)
    data = openpyxl.Workbook()
    print(data.get_sheet_names())
    sheet=data.get_sheet_by_name("Sheet")
    cell=""
    end = False
    pages = 0
    while True:
        try:
            nextPage = None
            nextPage = br.find_element_by_class_name("paging_next")
        except:
            end = True
        pages += 1
        tr = br.find_elements_by_tag_name("tr")
        head = tr[0].text.split(" ")
        if pages==1:
            for k in range (len(head)):
                cell=get_column_letter(k+1)+"1"
                sheet[cell]=head[k]
        for i in range(1, len(tr)):
            #meta = {}
            line = tr[i].text.split(" ")
            for j in range(len(line)):
                cell=get_column_letter(j+1)+str(int(line[0])+1)
                sheet[cell] = line[j]
            #data.append(meta)
            #print(meta)
        if end:
            br.close()
            break
        nextPage.click()
        time.sleep(0.2)
    #with open('data.txt', 'w') as f:
        #json.dump(data,f)
    data.save("data.xlsx")
    #print(pages)


if __name__ == "__main__":
    main()
